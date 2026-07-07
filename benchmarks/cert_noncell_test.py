#!/usr/bin/env python3
"""The one reachable false-certification class: a NON-CELL reference left unshifted.

xlq certify diffs the foreign edit against xlq's transform via diff::snapshot, which
reads only sheet CELLS. A foreign edit that correctly shifts every cell formula but
leaves a DEFINED-NAME refers-to unshifted would be invisible to that diff — a silent
false certification IF the gate doesn't fail-close it. This builds exactly that edit
and runs it through certify: it either CERTIFIES (a demonstrated hole -> we must fix)
or REFUSES via a residual (we state the boundary)."""
import json, os, re, shutil, subprocess, sys, zipfile
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

XLQ = "/home/soh/aix/xlq/target/release/xlq"
W = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/noncell"
K = 2


def build_orig(path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    for r in range(1, 15):
        ws[f"A{r}"] = float(r * 10)          # A10 = 100
    # a NORMAL defined name (not spelled like a cell) -> passes the aliasing gate
    wb.defined_names.add(DefinedName("Rate", attr_text="S!$A$10"))
    ws["C5"] = "=A10*2"                        # a plain cell formula (will shift)
    ws["C6"] = "=Rate*2"                        # uses the defined name (name target must move)
    for suf in ("", ".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        if os.path.exists(path + suf):
            os.remove(path + suf)
    wb.save(path)


def defined_name_target(path):
    data = zipfile.ZipFile(path).read("xl/workbook.xml").decode("utf-8", "replace")
    m = re.search(r'<definedName[^>]*name="Rate"[^>]*>([^<]*)</definedName>', data)
    return m.group(1) if m else None


def xlq_restructure(src, dst):
    shutil.copy(src, dst)
    for suf in (".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        if os.path.exists(dst + suf):
            os.remove(dst + suf)
    r = subprocess.run([XLQ, "restructure", dst, "--sheet", "S", "--op", "insert-rows",
                        "--at", str(K), "--count", "1", "--actor", "t"], capture_output=True, text=True)
    return '"rev"' in r.stdout, r.stdout


def revert_defined_name(xlq_out, dst, orig_target):
    """xlq's correct output, but the defined name reverted to its UNSHIFTED target —
    every sheet cell correctly shifted, only the non-cell reference wrong."""
    names = zipfile.ZipFile(xlq_out).namelist()
    buf = {n: zipfile.ZipFile(xlq_out).read(n) for n in names}
    wbxml = buf["xl/workbook.xml"].decode("utf-8", "replace")
    wbxml2 = re.sub(r'(<definedName[^>]*name="Rate"[^>]*>)[^<]*(</definedName>)',
                    r'\g<1>' + orig_target + r'\g<2>', wbxml, count=1)
    buf["xl/workbook.xml"] = wbxml2.encode("utf-8")
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
        for n in names:
            z.writestr(n, buf[n])


def certify(src, edited):
    r = subprocess.run([XLQ, "certify", src, edited, "--sheet", "S", "--op", "insert-rows",
                        "--at", str(K), "--count", "1"], capture_output=True, text=True)
    try:
        return json.loads(r.stdout)
    except Exception:
        return {"status": "PARSE_ERROR", "raw": r.stdout[:200]}


if __name__ == "__main__":
    os.makedirs(W, exist_ok=True)
    orig = f"{W}/orig.xlsx"; build_orig(orig)
    print("original defined name 'Rate' ->", defined_name_target(orig))
    xo = f"{W}/xlq.xlsx"; ok, out = xlq_restructure(orig, xo)
    if not ok:
        print("xlq restructure did not apply:", out[:200]); sys.exit(2)
    print("xlq-shifted defined name 'Rate' ->", defined_name_target(xo), "(should be $A$11)")

    # sanity: certify xlq's OWN correct output -> CERTIFIED
    c_ok = certify(orig, xo)
    print("certify(xlq's own correct edit) ->", c_ok.get("status"))

    # the attack: sheet cells correctly shifted, defined name reverted to $A$10
    foreign = f"{W}/foreign.xlsx"
    revert_defined_name(xo, foreign, "S!$A$10")
    print("foreign edit defined name 'Rate' ->", defined_name_target(foreign), "(unshifted = WRONG)")
    c_bad = certify(orig, foreign)
    print("certify(foreign: cells shifted, defined name NOT shifted) ->", c_bad.get("status"))
    print("  diff_counts:", c_bad.get("diff_counts"))

    if c_bad.get("status") == "CERTIFIED":
        print("\nRESULT: FALSE CERTIFICATION — certify blessed a non-cell-reference corruption "
              "(diff::snapshot does not compare defined names). REAL HOLE.")
        sys.exit(1)
    else:
        print("\nRESULT: certify REFUSED the non-cell corruption -", c_bad.get("reason", "")[:120])
        sys.exit(0)
