#!/usr/bin/env python3
"""Real-corpus edit-path A/B across ops (Task 1 scale on real data).

The headline A/B (86.6%) was insert-row@2 on real workbooks vs the Excel cache. This
runs REAL workbooks across insert-row AND insert-col, adjudicated by LibreOffice
SELF-CONSISTENT value-preservation (recompute the original with LibreOffice as the
'before', so there is no Excel-vs-LibreOffice disagreement and no reliability gate —
ACCRINT/BESSEL are measurable). A blank row/col insert is value-preserving under a
correct shift; a formula whose recomputed value at its shifted position changes = a
silent corruption. openpyxl shifts no references; xlq shifts via the Z3-backed algebra.
Delete ops are covered cleanly on generated workbooks (agent_ab_v2); real-corpus delete
value-preservation is confounded by range-clamp, so we keep the real sweep to inserts."""
import glob, os, re, shutil, subprocess, sys, zipfile
sys.path.insert(0, os.path.dirname(__file__))
from forward_correctness import (XLQ, first_sheet_part, col_num, VOLATILE, lo_grid, val_at)

WORK = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/abreal"
CORPUS = sorted(glob.glob("/home/soh/aix/vendor/upstream/xlsx/tests/**/*.xlsx", recursive=True))
CELL = re.compile(rb'<c r="([A-Z]+)(\d+)"((?:(?!</c>).)*?<f[ >](?:(?!</c>).)*?)</c>', re.S)


def strip_caches(src, dst):
    """Copy `src` to `dst` with formula caches (<v> inside formula cells) removed and
    fullCalcOnLoad set, so LibreOffice RECOMPUTES every formula rather than reading the
    Excel-authored cache. Needed so 'before' is LibreOffice-computed (like the edited
    files), not Excel-computed — otherwise functions where LibreOffice != Excel
    (ACCRINT/TIME/DAYS360) show a spurious before/after divergence unrelated to the shift."""
    names = zipfile.ZipFile(src).namelist()
    buf = {n: zipfile.ZipFile(src).read(n) for n in names}
    for n in list(buf):
        if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
            t = buf[n].decode("utf-8", "replace")
            # drop the <v>..</v> that follows an <f> within a cell
            t = re.sub(r'(</f>)<v>[^<]*</v>', r'\1', t)
            buf[n] = t.encode("utf-8")
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
        for n in names:
            z.writestr(n, buf[n])
    return dst


def zip_first_sheet_name(path):
    try:
        d = zipfile.ZipFile(path).read("xl/workbook.xml").decode("utf-8", "replace")
        m = re.search(r'<sheet\b[^>]*\bname="([^"]*)"', d)
        return m.group(1) if m else None
    except Exception:
        return None


def formula_cells(path):
    """[(col_num,row)] of first-sheet formula cells (non-volatile)."""
    part = first_sheet_part(path)
    if not part:
        return []
    data = zipfile.ZipFile(path).read(part)
    out = []
    for m in CELL.finditer(data):
        if VOLATILE.search(m.group(3)):
            continue
        out.append((col_num(m.group(1).decode()), int(m.group(2))))
    return out


def xlq_edit(src, sheet, op, at, work):
    dst = os.path.join(work, "xlq.xlsx"); shutil.copy(src, dst)
    for suf in (".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        if os.path.exists(dst + suf):
            os.remove(dst + suf)
    r = subprocess.run([XLQ, "restructure", dst, "--sheet", sheet, "--op", op, "--at", str(at),
                        "--count", "1", "--actor", "r"], capture_output=True, text=True)
    return dst if '"rev"' in r.stdout else None


def opx_edit(src, op, at, work):
    import openpyxl
    dst = os.path.join(work, "opx.xlsx")
    wb = openpyxl.load_workbook(src)
    ws = wb[wb.sheetnames[0]]
    if op == "insert-rows":
        ws.insert_rows(at, 1)
    else:
        ws.insert_cols(at, 1)
    wb.save(dst)
    return dst


def corrupted(before_grid, edited_path, cells, op, at, work, detail=None):
    """True if any formula cell's value diverges from before at its shifted position.
    If `detail` is a list, append (col,row,v0,v1) of the first divergence."""
    from openpyxl.utils import get_column_letter
    g = lo_grid(edited_path, work)
    if g is None:
        return None
    checked = 0
    for c, r in cells:
        nc = c + 1 if (op == "insert-cols" and c >= at) else c
        nr = r + 1 if (op == "insert-rows" and r >= at) else r
        v0 = val_at(before_grid, c, r); v1 = val_at(g, nc, nr)
        if v0 is None or v1 is None:
            continue
        checked += 1
        if abs(v1 - v0) > 1e-6 * max(abs(v0), abs(v1), 1.0):
            if detail is not None:
                detail.append(f"{get_column_letter(c)}{r}->{get_column_letter(nc)}{nr}: {v0} -> {v1}")
            return True
    return False if checked else None


if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 24
    from collections import Counter
    tally = {op: Counter() for op in ("insert-rows", "insert-cols")}
    done = 0
    for p in CORPUS:
        if done >= limit:
            break
        sheet = zip_first_sheet_name(p)
        cells = formula_cells(p)
        if not sheet or len(cells) < 2:
            continue
        work = os.path.join(WORK, str(done)); os.makedirs(work, exist_ok=True)
        stripped = strip_caches(p, os.path.join(work, "orig_nocache.xlsx"))
        before = lo_grid(stripped, work)                # LibreOffice-recomputed 'before'
        if before is None:
            shutil.rmtree(work, ignore_errors=True); continue
        used = False
        for op, at in [("insert-rows", 2), ("insert-cols", 2)]:
            xf = xlq_edit(p, sheet, op, at, work)
            if xf is None:
                tally[op]["xlq_refused"] += 1
            else:
                det = []
                xc = corrupted(before, xf, cells, op, at, work, det)
                if xc:
                    print(f"    !! xlq CORRUPT {op} {os.path.basename(p)}: {det[0] if det else '?'}", flush=True)
                if xc is not None:
                    tally[op]["xlq_corrupt" if xc else "xlq_faithful"] += 1; used = True
            try:
                of = opx_edit(p, op, at, work)
                oc = corrupted(before, of, cells, op, at, work)
                if oc is not None:
                    tally[op]["opx_corrupt" if oc else "opx_faithful"] += 1
            except Exception:
                pass
        shutil.rmtree(work, ignore_errors=True)
        if used:
            done += 1
            print(f"  [{done}] {os.path.relpath(p, '/home/soh/aix/vendor/upstream/xlsx/tests')[:40]}", flush=True)

    import json
    summary = {"benchmark": "real-corpus edit-path A/B (LibreOffice self-consistent), insert-row + insert-col",
               "workbooks": done, "per_op": {}}
    for op in tally:
        t = tally[op]
        oc, of = t["opx_corrupt"], t["opx_faithful"]
        xc, xf, xr = t["xlq_corrupt"], t["xlq_faithful"], t["xlq_refused"]
        summary["per_op"][op] = {
            "openpyxl_corrupt": oc, "openpyxl_faithful": of,
            "openpyxl_silent_corruption_rate": round(oc / (oc + of), 3) if (oc + of) else None,
            "xlq_corrupt": xc, "xlq_faithful": xf, "xlq_refused": xr,
            "xlq_false_faithful_rate": round(xc / (xc + xf), 3) if (xc + xf) else None,
        }
    json.dump(summary, open("/home/soh/aix/benchmarks/agent_ab_real.json", "w"), indent=2)
    print("\n" + json.dumps(summary["per_op"], indent=2))
