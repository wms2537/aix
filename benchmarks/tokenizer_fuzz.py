#!/usr/bin/env python3
"""Differential fuzzer for xlq's reference-shift tokenizer (the TCB's ref predicate).

Per adversarial review: replace "N bugs fixed" with "validated vs an independent
tokenizer." A token is a cell ref IFF col_to_num(col) in 1..=16384 AND row in
1..=1048576, delimited by ref-boundary chars (not preceded/followed by an
identifier char, not followed by '(' — a function call). We generate many formulas
mixing Excel function names (esp. digit-bearing: LOG10, ATAN2, BIN2DEC, SUMX2MY2,
IMLOG2, T.DIST...), real cell refs, out-of-grid tokens, and defined names; shift
each with a SIMPLE independent reference impl; and compare against xlq restructure.
Any disagreement is a tokenizer bug in one of them."""
import os, re, subprocess, sys

XLQ = "/home/soh/aix/xlq/target/release/xlq"
WORK = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/fuzz"
K, COUNT = 2, 1
BASE_ROW = 60          # place test formulas far below the insert so refs shift uniformly

# Excel functions whose NAMES embed or trail digits — the corruption class.
DIGIT_FUNCS = ["LOG10", "ATAN2", "SUMX2MY2", "SUMX2PY2", "SUMXMY2", "BIN2DEC", "BIN2HEX",
               "BIN2OCT", "DEC2BIN", "DEC2HEX", "DEC2OCT", "HEX2BIN", "HEX2DEC", "HEX2OCT",
               "OCT2BIN", "OCT2DEC", "OCT2HEX", "IMLOG2", "IMLOG10"]
PLAIN_FUNCS = ["SUM", "LOG", "SIN", "COS", "IF", "ROUND", "ABS"]
# Tokens that are DECIDABLY not cell references (out of the A1..XFD1048576 grid), so
# a correct shifter must leave them alone. NOTE: grid-VALID name-lookalikes (Q1,
# FY2021, Tax2020 — syntactically identical to real cells) are deliberately excluded:
# distinguishing a defined name from a cell of the same spelling is UNDECIDABLE from
# the formula text alone (it needs the workbook's defined-names table) — a documented
# limitation, not a tokenizer bug. This fuzzer validates the DECIDABLE predicate.
NAMES = ["Sales2020", "XFE9", "XFF1", "ZZZ9", "A2000000", "AAAA5", "ZZ99999999"]


def col_num(s):
    n = 0
    for c in s.upper():
        n = n * 26 + (ord(c) - 64)
    return n


_CELL = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)([0-9]+)")


def ref_shift(formula, op, k, count):
    """Independent, obviously-correct A1 shifter for a row insert OR delete at k.
    A token is a reference iff it is GRID-VALID (col in A..XFD, row in 1..1048576),
    boundary-delimited, and not a function call. insert: row>=k -> +count. delete:
    row in [k,k+count) -> #REF!; row>=k+count -> -count; else unchanged."""
    out, i, n = [], 0, len(formula)
    while i < n:
        ch = formula[i]
        if ch == '"':                                  # copy string literal verbatim
            j = i + 1
            while j < n and formula[j] != '"':
                j += 1
            out.append(formula[i:j + 1]); i = j + 1; continue
        m = _CELL.match(formula, i)
        if m:
            prev = formula[i - 1] if i > 0 else ""
            nxt = formula[m.end()] if m.end() < n else ""
            delim_ok = not (prev.isalnum() or prev in ("_", ".", "$", "!", "'"))
            tail_ok = not (nxt.isalpha() or nxt in ("_", "("))    # tuple: avoids "" in "_(" == True
            grid_ok = 1 <= col_num(m.group(2)) <= 16384 and 1 <= int(m.group(4)) <= 1048576
            if delim_ok and tail_ok and grid_ok:
                col, row = m.group(2), int(m.group(4))
                if op == "insert-rows":
                    nr = row + count if row >= k else row
                    out.append(f"{m.group(1)}{col}{m.group(3)}{nr}")
                else:                                   # delete-rows
                    if k <= row < k + count:
                        out.append("#REF!")
                    else:
                        nr = row - count if row >= k + count else row
                        out.append(f"{m.group(1)}{col}{m.group(3)}{nr}")
                i = m.end(); continue
        out.append(ch); i += 1
    return "".join(out)


def gen_formulas():
    """A broad set of formulas exercising the corruption class."""
    fs = []
    for fn in DIGIT_FUNCS:
        fs.append(f"={fn}(A5)")
        fs.append(f"={fn}(A5,B5)")
        fs.append(f"=A5+{fn}(A9)*B1")
    for fn in PLAIN_FUNCS:
        fs.append(f"={fn}(A5:A9)")
        fs.append(f"={fn}(A5,$B$9,C1)")
    for nm in NAMES:
        fs.append(f"={nm}+A5")
        fs.append(f"={nm}*B9-A1")
    fs += ["=A5+A9+A1", "=$A$5+B$9+$C1", "=XFD5+A9", "=A1048576+A5",
           '=IF(A5>0,"A5",B9)', "=A5:A9", "=SUM(A5:A9)+LOG10(B9)"]
    return fs


def run_op(op, at, count):
    """Place every generated formula at BASE_ROW, apply the op, and compare each
    result to the independent shifter. Returns (n, mismatches)."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    fs = gen_formulas()
    # the independent oracle models single-cell shifting, not the 6-case range clamp
    # (xlq's clamp arithmetic is validated by its own unit tests); so under delete we
    # fuzz single-cell formulas only — the tokenizer (ref identification) is identical
    # for both, which is what this fuzzer validates.
    if op == "delete-rows":
        fs = [f for f in fs if ":" not in f]
    wb = openpyxl.Workbook(); ws = wb.active
    for r in (1, 5, 9):
        for c in "ABCD":
            ws[f"{c}{r}"] = 1
    placed = []
    for idx, f in enumerate(fs):
        col = get_column_letter(idx + 1)
        ws[f"{col}{BASE_ROW}"] = f
        placed.append((col, f))
    path = os.path.join(WORK, f"fuzz_{op}.xlsx")
    for suf in ("", ".xlq.jsonl", ".rev-1.xlsx", ".rev-2.xlsx", ".rev-3.xlsx", ".xlq.lock"):
        if os.path.exists(path + suf):
            os.remove(path + suf)
    wb.save(path)
    r = subprocess.run([XLQ, "restructure", path, "--sheet", ws.title, "--op", op,
                        "--at", str(at), "--count", str(count), "--actor", "fuzz"],
                       capture_output=True, text=True)
    if '"rev"' not in r.stdout:
        print(f"restructure {op} did not apply:", (r.stdout or r.stderr)[:300]); sys.exit(2)
    ws2 = openpyxl.load_workbook(path)[ws.title]
    new_row = BASE_ROW + (count if op == "insert-rows" else -count)  # formula's new position
    mism = []
    for col, f in placed:
        expected = ref_shift(f, op, at, count)
        got = ws2[f"{col}{new_row}"].value
        got = got if got is not None else ""
        if str(got) != str(expected):
            mism.append({"op": op, "formula": f, "expected": expected, "xlq": got})
    return len(placed), mism


def main():
    import json
    os.makedirs(WORK, exist_ok=True)
    total, all_mism = 0, []
    for op, at, count in [("insert-rows", 2, 1), ("delete-rows", 5, 1)]:
        n, mism = run_op(op, at, count)
        total += n; all_mism += mism
        print(f"  {op}@{at}: {n} formulas, {len(mism)} disagreements")
    print(f"differential fuzz TOTAL: {total} (formula,op) pairs, {len(all_mism)} disagreements "
          f"with the independent grid-validity tokenizer")
    for m in all_mism[:20]:
        print(f"  [{m['op']}] IN {m['formula']}\n   exp {m['expected']}\n   xlq {m['xlq']}")
    json.dump({"pairs_tested": total, "disagreements": len(all_mism), "cases": all_mism},
              open("/home/soh/aix/benchmarks/tokenizer_fuzz.json", "w"), indent=2)
    sys.exit(1 if all_mism else 0)


if __name__ == "__main__":
    main()
