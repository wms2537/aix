#!/usr/bin/env python3
"""Task prep for the guarded-vs-unguarded LIVE-AGENT STUDY (op = insert-row@2).

Selects real corpus workbooks on which ALL THREE instruments of the study can
rule cleanly — the guard (foreign_certify), the ground truth (ref_shift), and
the artifact builder (openpyxl). A workbook is a task only if:

  - every first-sheet formula is fully A1-modeled (uncertifiable_formula False
    for all): no tables, cross-sheet refs, whole-row/col refs, defined names —
    otherwise the guard would trivially refuse and the task measures nothing;
  - no volatile / position-dependent functions (OFFSET, INDIRECT, ROW, ...):
    foreign_certify.extract deliberately SKIPS those cells (their value
    legitimately changes under insert), so an agent error there is invisible
    to the guard — including them would blur the false-cert metric, so they
    are excluded from the study universe up front (counted);
  - no shared-formula follower cells (self-closing or empty <f/>): their
    formula text cannot be presented to the agent, and openpyxl expands them
    on save, injecting graph nodes the guard cannot account for;
  - the "first sheet" is the SAME sheet for every instrument: workbook.xml
    order (task sheet name) == lowest-numbered sheet part (guard + truth
    extraction) == openpyxl sheetnames[0] (artifact builder);
  - openpyxl can load the file (score.py builds the artifact with it);
  - at least one formula references a row >= 2, so the insert at row 2 is a
    NON-TRIVIAL reference-shift task;
  - at least one cell is INSIDE the ground truth's grammar (ref_shift) AND its
    formula must change under the insert — otherwise the task could never be
    truth-visibly wrong. NOTE ref_shift's grammar excludes every formula that
    contains a RANGE (its RANGE_FN gate matches all ranges, not just
    function-endpoint ones), so this requirement bites hard;
  - formula-cell count within [CELL_LO, CELL_HI] (env, default 2..40);
  - not a byte-identical duplicate of an already-selected workbook.

Output: tasks.json — live3way's compact per-task format (file, sheet, k,
cells=[{cell,row,col,formula,cached_value}]) PLUS a difficulty tag
(n_formulas, has_absolute_refs, has_ranges). Formula text is XML-unescaped
(human formula space: '<' not '&lt;'); score.py re-escapes when splicing.

usage: prep.py [N]        (default N=30)
"""
import glob, hashlib, json, os, re, sys, zipfile
from collections import Counter

BENCH = "/home/soh/aix/benchmarks"
sys.path.insert(0, BENCH)
from foreign_certify import uncertifiable_formula, first_sheet_part, FTAG   # noqa: E402
from forward_correctness import VOLATILE                                     # noqa: E402
from live3way_prep import zip_first_sheet_name, formula_cells, refs_below_k, K  # noqa: E402
from score import sheet_part_by_name                                         # noqa: E402
from shift_correctness_real import RANGETOK, ref_shift, norm                 # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
CORPUS_DIR = "/home/soh/aix/vendor/upstream/xlsx/tests"
CORPUS = sorted(glob.glob(CORPUS_DIR + "/**/*.xlsx", recursive=True))
F_SELFCLOSED = re.compile(rb"<f[^>]*/>")
_STR = re.compile(r'"[^"]*"')


def _unesc(s):
    """XML entity -> character (the 5 standard entities; &amp; last)."""
    return (s.replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'").replace("&amp;", "&"))


def skip_reason(path, sheet, lo, hi):
    """None if the workbook qualifies as a task, else a short skip reason.
    NO GUESSING: any condition we cannot positively verify -> skip."""
    part = first_sheet_part(path)
    if not part:
        return "no_sheet_part"
    try:
        data = zipfile.ZipFile(path).read(part)
    except Exception:
        return "unreadable_zip"
    # all three instruments must agree on which sheet is "the first sheet"
    try:
        if sheet_part_by_name(zipfile.ZipFile(path), sheet) != part:
            return "sheet_part_mismatch"
    except Exception:
        return "sheet_part_mismatch"
    if F_SELFCLOSED.search(data):
        return "shared_formula_followers"
    fs = [m.group(1).decode("utf-8", "replace") for m in FTAG.finditer(data)]
    if not fs:
        return "no_formulas"
    if any(not f.strip() for f in fs):
        return "empty_formula_body"
    if any(uncertifiable_formula(f) for f in fs):
        return "uncertifiable_formula"        # guard would trivially refuse
    if any(VOLATILE.search(f.encode("utf-8", "replace")) for f in fs):
        return "volatile_function"            # guard is blind to these cells
    cells = formula_cells(path)
    if not (lo <= len(cells) <= hi):
        return "size_band"
    if not refs_below_k(cells, K):
        return "trivial_no_row_ge_k_refs"
    # the GROUND TRUTH (ref_shift) must be able to rule on the task: at least one
    # cell inside its grammar WHOSE FORMULA MUST CHANGE under the insert — else a
    # botched task could never be truth-visibly wrong (note: ref_shift's grammar
    # excludes ALL formulas containing ranges, so this bites hard).
    visible_shift = False
    for _, _, _, f, _ in cells:
        e = ref_shift(_unesc(f), "row", "insert-rows", K, 1)
        if e is not None and norm(e) != norm(_unesc(f)):
            visible_shift = True
            break
    if not visible_shift:
        return "no_truth_visible_shift"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        if not wb.sheetnames or wb.sheetnames[0] != sheet:
            return "openpyxl_first_sheet_mismatch"
    except Exception:
        return "openpyxl_load_failed"
    return None


def difficulty(cells):
    stripped = [_STR.sub("", f) for _, _, _, f, _ in cells]
    evaluable = shift_cells = 0
    for _, _, _, f, _ in cells:
        e = ref_shift(_unesc(f), "row", "insert-rows", K, 1)
        if e is None:
            continue
        evaluable += 1
        if norm(e) != norm(_unesc(f)):
            shift_cells += 1
    return {"n_formulas": len(cells),
            "has_absolute_refs": any("$" in s for s in stripped),
            "has_ranges": any(RANGETOK.search(s) for s in stripped),
            "truth_evaluable_cells": evaluable,   # cells inside ref_shift's grammar
            "truth_shift_cells": shift_cells,     # evaluable cells whose formula must change
            "truth_total": evaluable == len(cells)}


if __name__ == "__main__":
    want = int(sys.argv[1]) if len(sys.argv) > 1 else 30
    lo = int(os.environ.get("CELL_LO", "2"))
    hi = int(os.environ.get("CELL_HI", "40"))
    tasks, reasons, seen = [], Counter(), set()
    for p in CORPUS:
        if len(tasks) >= want:
            break
        h = hashlib.md5(open(p, "rb").read()).hexdigest()
        if h in seen:
            reasons["duplicate_content"] += 1
            continue
        seen.add(h)
        sheet = zip_first_sheet_name(p)
        if not sheet:
            reasons["no_sheet_name"] += 1
            continue
        r = skip_reason(p, sheet, lo, hi)
        if r:
            reasons[r] += 1
            continue
        cells = formula_cells(p)
        rel = os.path.relpath(p, CORPUS_DIR)
        tasks.append({
            "file": rel, "sheet": sheet, "k": K,
            "difficulty": difficulty(cells),
            "cells": [{"cell": a1, "row": rr, "col": c,
                       "formula": _unesc(f), "cached_value": v}
                      for a1, rr, c, f, v in cells],
        })
    with open(os.path.join(HERE, "tasks.json"), "w") as f:
        json.dump(tasks, f, indent=2)
    print(f"selected {len(tasks)}/{want} tasks (band {lo}..{hi} formula cells) "
          f"from {len(CORPUS)} corpus files")
    for t in tasks:
        d = t["difficulty"]
        print(f"  {t['file'][:52]:52} n={d['n_formulas']:3} "
              f"abs={'y' if d['has_absolute_refs'] else 'n'} "
              f"rng={'y' if d['has_ranges'] else 'n'}")
    print("skip reasons:", dict(sorted(reasons.items(), key=lambda kv: -kv[1])))
