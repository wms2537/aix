#!/usr/bin/env python3
"""E-structural. Insert one row via three tools and score three axes:

  (1) CORRECTNESS  — fraction of references shifted to their documented target.
                     A structural edit MUST rewrite reference-bearing parts, so
                     this, not byte-identity, is the headline.
  (2) MINIMAL-PATCH — for the correct tool, the parts it changed differ ONLY in
                     reference coordinates; every non-reference part is byte-
                     identical.
  (3) RECOMPUTE     — the edited file recomputes to the same values as before
                     for formulas whose logical target is unchanged.

Contrast: xlq shifts every reference and touches only coordinate bytes; openpyxl
shifts NONE (silently wrong values); LibreOffice shifts by engine but rewrites
the whole container (round-trip fidelity loss)."""
import json, os, re, shutil, subprocess, zipfile

XLQ = "/home/soh/aix/xlq/target/release/xlq"
FIX = "/home/soh/aix/fixtures/structural/full.xlsx"
SCRATCH = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/eval"
os.makedirs(SCRATCH, exist_ok=True)

# Expected post-edit state after inserting 1 row at row 5 on "Data".
# formula-by-logical-identity: (locate by the formula that USED to be at Xn)
EXPECT = {
    "SUM":       "=SUM(B2:B8)",   # B8 straddles insert-at-5 -> grows
    "double":    "=B6*2",         # was =B5*2, B5 -> B6
    "abs":       "=$B$9",         # was =$B$8
    "xref":      "=Data!B9",      # Report!A1 was =Data!B8
    "defined":   "Data!$B$9",     # Total was Data!$B$8
    "chart_ref": "Data!$B$2:$B$8" # chart data was Data!$B$2:$B$7 -> grows
}


def parts(path):
    z = zipfile.ZipFile(path)
    return {n: z.read(n) for n in z.namelist() if not n.endswith("/")}


def formulas_present(path):
    """Return the set of formula strings present anywhere in Data + the
    cross-sheet ref + defined name + chart ref, by logical identity."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    d = wb["Data"]
    allf = [c.value for row in d.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=")]
    got = {
        "SUM": next((f for f in allf if f.startswith("=SUM")), None),
        "double": next((f for f in allf if re.fullmatch(r"=B\d+\*2", f)), None),
        "abs": next((f for f in allf if f.startswith("=$B$")), None),
        "xref": wb["Report"]["A1"].value,
        "defined": next((v.attr_text for k, v in wb.defined_names.items()), None),
    }
    # chart ref (sheet name may be quoted 'Data'! — normalize the quotes away)
    try:
        z = zipfile.ZipFile(path)
        chart = z.read("xl/charts/chart1.xml").decode("utf8", "replace")
        m = re.findall(r"'?Data'?!\$B\$\d+:\$B\$\d+", chart)
        got["chart_ref"] = m[0].replace("'", "") if m else None
    except KeyError:
        got["chart_ref"] = None
    return got


def score_correctness(got):
    per = {k: (got.get(k) == v) for k, v in EXPECT.items()}
    return sum(per.values()), len(EXPECT), per


def byte_identical(orig, new):
    o, n = parts(orig), parts(new)
    same = [k for k in o if n.get(k) == o[k]]
    changed = [k for k in o if k in n and n[k] != o[k]]
    dropped = [k for k in o if k not in n]
    added = [k for k in n if k not in o]
    return same, changed, dropped, added


def minimal_patch_check(orig, new):
    """For each CHANGED part, confirm the only textual differences are numeric
    coordinate shifts inside references — i.e. the non-digit skeleton of the
    part is unchanged. A crude but honest proxy: strip all digits and compare;
    if the digit-stripped bytes are identical, every change was a coordinate."""
    o, n = parts(orig), parts(new)
    results = {}
    for k in o:
        if k not in n or n[k] == o[k]:
            continue
        if k.endswith(".xml"):
            ob = re.sub(rb"\d+", b"#", o[k])
            nb = re.sub(rb"\d+", b"#", n[k])
            # row/cell inserts add whole elements, so the edited sheet's skeleton
            # legitimately grows; for non-sheet parts (chart, workbook, xref
            # sheet) the skeleton must be identical (pure coordinate change).
            results[k] = {"digit_stripped_identical": ob == nb}
    return results


def arm_xlq():
    dst = f"{SCRATCH}/xlq.xlsx"
    # clear stale receipt journal / rev files from a prior run (else the chain
    # detects the fresh copy as an external edit and refuses to write)
    for suffix in (".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        p = dst + suffix
        if os.path.exists(p):
            os.remove(p)
    shutil.copy(FIX, dst)
    out = subprocess.run(
        [XLQ, "restructure", dst, "--sheet", "Data", "--op", "insert-rows",
         "--at", "5", "--count", "1", "--actor", "eval"],
        capture_output=True, text=True)
    rep = json.loads(out.stdout) if out.stdout.strip().startswith("{") else {"error": out.stderr}
    # RECOMPUTE (executed, not asserted): reopen the committed file in the
    # engine via `xlq calc` and confirm the SUM recomputes to its pre-edit value
    # (760); the inserted blank row contributes 0, data shifts, total unchanged.
    calc = subprocess.run([XLQ, "calc", dst], capture_output=True, text=True)
    recompute_sum = None
    try:
        cj = json.loads(calc.stdout)
        for c in cj.get("changed", []):
            if isinstance(c.get("formula"), str) and c["formula"].startswith("=SUM"):
                recompute_sum = c.get("recomputed")
    except Exception:
        pass
    got = formulas_present(dst)
    correct, tot, per = score_correctness(got)
    same, changed, dropped, added = byte_identical(FIX, dst)
    mp = minimal_patch_check(FIX, dst)
    # non-sheet changed parts must be pure-coordinate (digit-stripped identical)
    nonsheet_pure = all(
        v["digit_stripped_identical"]
        for k, v in mp.items()
        if "worksheets/sheet" not in k
    )
    return {
        "tool": "xlq restructure",
        "correctness": f"{correct}/{tot}", "per_ref": per,
        "recompute_sum": recompute_sum, "recompute_expected": "760",
        "recompute_ok": recompute_sum == "760",
        "committed": "rev" in rep, "residuals": rep.get("edit", {}).get("residuals"),
        "parts_byte_identical": len(same), "parts_total": len(same) + len(changed) + len(dropped),
        "parts_changed": changed, "parts_dropped": dropped,
        "minimal_patch_nonsheet_pure_coordinate": nonsheet_pure,
        "minimal_patch_detail": mp,
    }


def arm_openpyxl():
    import openpyxl
    dst = f"{SCRATCH}/openpyxl.xlsx"
    wb = openpyxl.load_workbook(FIX)
    wb["Data"].insert_rows(5, 1)
    wb.save(dst)
    got = formulas_present(dst)
    correct, tot, per = score_correctness(got)
    same, changed, dropped, added = byte_identical(FIX, dst)
    return {
        "tool": "openpyxl insert_rows",
        "correctness": f"{correct}/{tot}", "per_ref": per,
        "parts_byte_identical": len(same), "parts_total": len(same) + len(changed) + len(dropped),
        "note": "insert_rows moves cells but does NOT rewrite formula references — "
                "every formula is silently wrong (e.g. =B5*2 now reads the blank inserted row)",
    }


def arm_libreoffice():
    d = f"{SCRATCH}/lo"
    if os.path.exists(d):
        shutil.rmtree(d)
    os.makedirs(d)
    shutil.copy(FIX, f"{d}/in.xlsx")
    # A real round-trip: import then export to a DIFFERENT dir so we compare
    # LO's serialization against the original. (LO's insert-row needs a Basic
    # macro; reference-shift is correct-by-engine, so the measurable axis here
    # is how much of the container a mere load-save rewrites.)
    outd = f"{d}/out"
    os.makedirs(outd)
    r = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", outd,
         f"{d}/in.xlsx"],
        capture_output=True, text=True, timeout=120,
        env={**os.environ, "HOME": d})
    saved = f"{outd}/in.xlsx"
    if not os.path.exists(saved):
        return {"tool": "LibreOffice", "error": "convert failed", "stderr": r.stderr[:300]}
    same, changed, dropped, added = byte_identical(FIX, saved)
    return {
        "tool": "LibreOffice load-save round-trip",
        "parts_byte_identical": len(same),
        "parts_total": len(same) + len(changed) + len(dropped),
        "parts_changed": len(changed), "parts_dropped": len(dropped), "parts_added": len(added),
        "note": "reference-shift is correct-by-engine (not measured here); this is the fidelity "
                "cost of a mere round-trip — every rewritten part loses byte-provenance",
    }


if __name__ == "__main__":
    res = {
        "fixture": "full.xlsx — chart + straddling SUM + absolute + cross-sheet + defined-name + merged",
        "operation": "insert 1 row at row 5 on sheet Data",
        "expected": EXPECT,
        "arms": {},
    }
    for name, fn in [("xlq", arm_xlq), ("openpyxl", arm_openpyxl), ("libreoffice", arm_libreoffice)]:
        try:
            res["arms"][name] = fn()
        except Exception as e:
            res["arms"][name] = {"error": f"{type(e).__name__}: {e}"}
    print(json.dumps(res, indent=2))
