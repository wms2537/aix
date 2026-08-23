#!/usr/bin/env python3
"""Stratified real-corpus fidelity benchmark.

Reads benchmarks/real_feature_manifest.summary.json, selects numeric data cells
from the first worksheet without printing cell contents, applies the same logical
set-cell edit through xlq, openpyxl (default keep_vba=False), and a LibreOffice
re-save proxy, then records byte-level OOXML part survival per file.
"""
import argparse, json, os, shutil, subprocess, sys, tempfile, zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "benchmarks"))
from fidelity import compare, read_parts, sha256_file

OPENPYXL_DRIVER = r'''import sys, warnings, openpyxl
warnings.simplefilter("ignore")
src, out, sheet, cell, to = sys.argv[1:6]
to = float(to)
if to.is_integer(): to = int(to)
wb = openpyxl.load_workbook(src)
wb[sheet][cell] = to
wb.save(out)
print("OK")
'''

def first_editable_sheet(path):
    import openpyxl, warnings
    warnings.simplefilter("ignore")
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    for name in wb.sheetnames:
        ws = wb[name]
        picked = pick_cell(ws)
        if picked[0] is not None:
            title = ws.title
            wb.close()
            return title, picked
    wb.close()
    return None, (None, None)

def pick_cell(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.data_type == "n" and not c.value is None and isinstance(c.value, (int,float)) and c.coordinate != "A1":
                return c.coordinate, float(c.value)
    return None, None

def run_xlq(src, work, sheet, cell, value, xlq):
    dst = Path(work) / "xlq.xlsx"
    shutil.copy(src, dst)
    patch_path = str(dst) + ".patch.json"
    patch = {"base_hash": sha256_file(dst), "actor": "real-feature-bench", "clock": 1751500000000,
             "seed": 1, "ops": [{"type":"set_cell","sheet":sheet,"cell":cell,"value":value}]}
    patch_path_obj = Path(patch_path); patch_path_obj.write_text(json.dumps(patch))
    r = subprocess.run([str(xlq), "apply", str(dst), patch_path], capture_output=True, text=True, timeout=180)
    if r.returncode != 0 or not dst.exists():
        return None, {"returncode":r.returncode, "stderr":r.stderr[-2000:], "stdout":r.stdout[-2000:]}
    try:
        report = json.loads(r.stdout)
    except Exception:
        report = {}
    return dst, {"self_reported_fidelity": report.get("fidelity"), "result_hash": report.get("result_hash")}

def run_openpyxl(src, work, sheet, cell, value):
    out = Path(work) / "openpyxl.xlsx"
    driver = Path(work) / "_driver.py"; driver.write_text(OPENPYXL_DRIVER)
    r = subprocess.run([sys.executable, str(driver), str(src), str(out), sheet, cell, str(value)], capture_output=True, text=True, timeout=180)
    if r.returncode != 0 or not out.exists():
        return None, {"returncode":r.returncode, "stderr":r.stderr[-2000:]}
    return out, {}

def run_libreoffice(src, work, soffice):
    outdir = Path(work) / "lo"; outdir.mkdir()
    profile = Path(work) / "lo_profile"
    cmd = [str(soffice), "--headless", "--norestore", f"-env:UserInstallation=file://{profile}",
           "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", str(outdir), str(src)]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=240)
    stem = Path(src).stem + ".xlsx"
    out = outdir / stem
    if not out.exists():
        return None, {"returncode":r.returncode, "stderr":(r.stderr+r.stdout)[-2000:]}
    return out, {"filter":"re-save proxy"}

def loads_ironcalc(path, load_only):
    if not load_only.exists(): return None
    r = subprocess.run([str(load_only), str(path)], capture_output=True, timeout=120)
    return r.returncode == 0

def loads_soffice(path, soffice, work):
    verify = Path(work)/f"verify-{os.getpid()}-{id(path)}"; verify.mkdir()
    try:
        r = subprocess.run([str(soffice), "--headless", "--norestore",
                            f"-env:UserInstallation=file://{work}/lo_verify_profile",
                            "--convert-to","xlsx:Calc MS Excel 2007 XML","--outdir",str(verify),str(path)],
                           capture_output=True, timeout=180)
        return (verify/(Path(path).stem+".xlsx")).exists()
    except Exception:
        return False

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--summary", default=str(ROOT/"benchmarks/real_feature_manifest.summary.json"))
    ap.add_argument("--root", default=str(ROOT))
    ap.add_argument("--out", default=str(ROOT/"benchmarks/real_feature_fidelity.json"))
    ap.add_argument("--xlq", default=str(ROOT/"xlq/target/release/xlq"))
    ap.add_argument("--load-only", dest="load_only", default=str(ROOT/"xlq/target/release/load-only"))
    ap.add_argument("--soffice", default="/usr/bin/soffice")
    ap.add_argument("--limit", type=int, default=50)
    args = ap.parse_args()
    summary = json.loads(Path(args.summary).read_text())
    sample = summary["sample"][:args.limit]
    results=[]
    root=Path(args.root); xlq=Path(args.xlq); lo_bin=Path(args.soffice); load_only=Path(args.load_only)
    work_root=Path(tempfile.mkdtemp(prefix="real_feature_fidelity_"))
    print(f"running {len(sample)} sampled files", file=sys.stderr)
    for i,row in enumerate(sample):
        src=root/"data/inthewild"/row["path"]; work=work_root/f"{i:03d}"; work.mkdir(parents=True)
        entry={"corpus":row["corpus"],"path":row["path"],"stratum":row["stratum"],"original_features":{k:v for k,v in row.items() if k.startswith("has_")}}
        try:
            sheet, (cell, oldval) = first_editable_sheet(src)
            if not cell:
                entry.update({"skip":"no_numeric_data_cell"}); results.append(entry); continue
            entry["edit"]={"sheet":sheet,"cell":cell}
            tools={}
            for tool in ("xlq","openpyxl","libreoffice"):
                try:
                    if tool=="xlq": out,extra=run_xlq(src,work,sheet,cell,oldval,xlq)
                    elif tool=="openpyxl": out,extra=run_openpyxl(src,work,sheet,cell,oldval)
                    else: out,extra=run_libreoffice(src,work,lo_bin)
                    if out is None: tools[tool]={"error":extra}; continue
                    m=compare(read_parts(src),out)
                    m["output_ext"]=out.suffix; m["output_loads_in_ironcalc"]=loads_ironcalc(out,load_only)
                    m["output_loads_in_soffice"]=loads_soffice(out,lo_bin,work)
                    m.update(extra); tools[tool]=m
                except Exception as e:
                    tools[tool]={"error":f"{type(e).__name__}: {e}"}
            entry["tools"]=tools
            ident={t:m.get("parts_byte_identical") for t,m in tools.items() if "error" not in m}
            total=next((m.get("parts_total") for m in tools.values() if "error" not in m),None)
            print(f"[{i+1}/{len(sample)}] {row['stratum']:20} {ident}/{total}", file=sys.stderr)
        except Exception as e:
            entry["error"]=f"{type(e).__name__}: {e}"
        finally:
            shutil.rmtree(work, ignore_errors=True)
        results.append(entry)
        Path(args.out).write_text(json.dumps({"experiment":"stratified real-corpus feature preservation",
                                               "seed":summary.get("seed"),"manifest_summary":args.summary,
                                               "methodology":{"libreoffice":"re-save proxy; upper bound on churn"},
                                               "results":results}, indent=2))
    shutil.rmtree(work_root, ignore_errors=True)

if __name__=="__main__": main()
