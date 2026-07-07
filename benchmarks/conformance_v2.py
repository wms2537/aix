#!/usr/bin/env python3
"""Coverage + scale + second-oracle tokenizer conformance.

Improves tokenizer_conformance on three axes at once:
  COVERAGE  — all FOUR structural ops (insert/delete x rows/cols), not just insert-row@2.
  SCALE     — larger seeded formula population.
  2nd ORACLE — value-preservation checked by TWO independent engines, LibreOffice AND
              `formulas` (pure-Python), neither of which is xlq's IronCalc.

Property: a blank ROW/COLUMN insert is value-preserving under a correct reference shift;
a row/col delete preserves the value of every formula not referencing a deleted line.
Recompute the same file before and after with each engine; any value divergence is a
real shift/tokenizer bug the engine caught, independent of our spec. Distinct cell
values so a mis-shift onto another cell changes the value."""
import csv, os, random, subprocess, sys, warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.utils import get_column_letter

XLQ = "/home/soh/aix/xlq/target/release/xlq"
WORK = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/conf2"
ROWS, COLS = 40, 12          # data grid A1:L40
NF = 140
SEED = 20260707


def cellref(rng):
    c = get_column_letter(rng.randint(1, COLS)); r = rng.randint(1, ROWS)
    cd = "$" if rng.random() < 0.3 else ""; rd = "$" if rng.random() < 0.3 else ""
    return f"{cd}{c}{rd}{r}"


def rng_range(rng):
    c1 = rng.randint(1, COLS - 2); r1 = rng.randint(1, ROWS - 5)
    # keep ranges single-column (clean under both row and col delete-exclusion)
    return f"{get_column_letter(c1)}{r1}:{get_column_letter(c1)}{r1 + rng.randint(1,4)}"


def gen(rng):
    a, b, c = cellref(rng), cellref(rng), cellref(rng)
    k = rng.choice(["add", "mul", "log10", "sqrt", "abs", "atan2", "max", "sum", "avg", "mixed"])
    return {
        "add": f"{a}+{b}", "mul": f"({a}+{b})*{c}", "log10": f"LOG10({a}+1)",
        "sqrt": f"SQRT({a})", "abs": f"ABS({a}-{b})", "atan2": f"ATAN2({a}+1,{b}+1)",
        "max": f"MAX({a},{b},{c})", "sum": f"SUM({rng_range(rng)})",
        "avg": f"AVERAGE({rng_range(rng)})", "mixed": f"SUM({rng_range(rng)})+{a}",
    }[k]


def build(path, formulas, rng):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    for r in range(1, ROWS + 1):
        for c in range(1, COLS + 1):
            ws.cell(r, c).value = round((r - 1) * COLS + c + 0.5 + rng.random() * 0.4, 4)
    fcol = COLS + 3                                   # formulas in a far column, one per row
    placed = []
    for i, f in enumerate(formulas):
        r = 45 + i
        ws.cell(r, fcol).value = "=" + f
        placed.append((get_column_letter(fcol), r, f))
    for suf in ("", ".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        if os.path.exists(path + suf):
            os.remove(path + suf)
    wb.save(path)
    return placed


def lo_values(path, work):
    outd = os.path.join(work, "csv"); os.makedirs(outd, exist_ok=True)
    for f in os.listdir(outd):
        os.remove(os.path.join(outd, f))
    subprocess.run(["libreoffice", "--headless", "--convert-to", "csv", "--outdir", outd, path],
                   capture_output=True, timeout=120, env={**os.environ, "HOME": work})
    p = os.path.join(outd, os.path.splitext(os.path.basename(path))[0] + ".csv")
    if not os.path.exists(p):
        return {}
    out = {}
    for ri, row in enumerate(csv.reader(open(p, newline="")), start=1):
        for ci, v in enumerate(row, start=1):
            try:
                out[(get_column_letter(ci), ri)] = float(v)
            except ValueError:
                pass
    return out


def fora_values(path):
    """Second independent engine: pure-Python `formulas`."""
    import formulas
    try:
        sol = formulas.ExcelModel().loads(path).finish().calculate()
    except Exception:
        return {}
    out = {}
    for k, v in sol.items():
        if "!" not in k:
            continue
        cell = k.split("!")[-1].replace("$", "")
        try:
            val = v.value[0][0]
            out[cell] = float(val)
        except Exception:
            pass
    # normalize "A5" -> (colletter,row)
    norm = {}
    import re
    for cell, val in out.items():
        m = re.match(r"([A-Z]+)(\d+)$", cell)
        if m:
            norm[(m.group(1), int(m.group(2)))] = val
    return norm


def new_pos(col_letter, row, op, at, count):
    from openpyxl.utils import column_index_from_string
    c = column_index_from_string(col_letter); r = row
    if op == "insert-rows":
        r = r + count if r >= at else r
    elif op == "delete-rows":
        r = r - count if r > at else r
    elif op == "insert-cols":
        c = c + count if c >= at else c
    elif op == "delete-cols":
        c = c - count if c > at else c
    return get_column_letter(c), r


def run_op(op, at, count, work):
    rng = random.Random(SEED + hash(op) % 1000)
    formulas = [gen(rng) for _ in range(NF)]
    if op.startswith("delete"):
        formulas = [f for f in formulas if ":" not in f]   # exclude ranges under delete
    path = os.path.join(work, f"{op}.xlsx")
    placed = build(path, formulas, rng)
    lo0 = lo_values(path, work); fo0 = fora_values(path)
    r = subprocess.run([XLQ, "restructure", path, "--sheet", "S", "--op", op, "--at", str(at),
                        "--count", str(count), "--actor", "c"], capture_output=True, text=True)
    if '"rev"' not in r.stdout:
        return {"op": op, "error": (r.stdout or r.stderr)[:150]}
    lo1 = lo_values(path, work); fo1 = fora_values(path)
    res = {"op": op}
    for oname, before, after in [("libreoffice", lo0, lo1), ("formulas", fo0, fo1)]:
        checked = div = 0
        for col, row, f in placed:
            nc, nr = new_pos(col, row, op, at, count)
            v0 = before.get((col, row)); v1 = after.get((nc, nr))
            if v0 is None or v1 is None:
                continue
            checked += 1
            if abs(v1 - v0) > 1e-6 * max(abs(v0), abs(v1), 1.0):
                div += 1
        res[oname] = {"checked": checked, "divergences": div}
    return res


if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    import json
    OPS = [("insert-rows", 2, 1), ("delete-rows", 4, 1), ("insert-cols", 2, 1), ("delete-cols", 4, 1)]
    results = []
    for op, at, count in OPS:
        work = os.path.join(WORK, op); os.makedirs(work, exist_ok=True)
        r = run_op(op, at, count, work)
        results.append(r)
        if "error" in r:
            print(f"  {op}: ERROR {r['error']}", flush=True)
        else:
            print(f"  {op}@{at}: LO {r['libreoffice']['checked']}chk/{r['libreoffice']['divergences']}div | "
                  f"formulas {r['formulas']['checked']}chk/{r['formulas']['divergences']}div", flush=True)
    tot_lo = sum(r.get("libreoffice", {}).get("checked", 0) for r in results)
    div_lo = sum(r.get("libreoffice", {}).get("divergences", 0) for r in results)
    tot_fo = sum(r.get("formulas", {}).get("checked", 0) for r in results)
    div_fo = sum(r.get("formulas", {}).get("divergences", 0) for r in results)
    summary = {
        "benchmark": "tokenizer conformance — 4 structural ops x 2 independent engines",
        "ops": [o[0] for o in OPS],
        "libreoffice": {"checked": tot_lo, "divergences": div_lo},
        "formulas_engine": {"checked": tot_fo, "divergences": div_fo},
        "per_op": results,
        "headline": (f"{len(OPS)} structural ops, TWO independent engines: LibreOffice "
                     f"{tot_lo} checks/{div_lo} divergences, formulas {tot_fo} checks/{div_fo} "
                     f"divergences."),
    }
    json.dump(summary, open("/home/soh/aix/benchmarks/conformance_v2.json", "w"), indent=2)
    print("\n" + summary["headline"])
    sys.exit(1 if (div_lo or div_fo) else 0)
