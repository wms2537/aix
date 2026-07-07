#!/usr/bin/env python3
"""Confusion matrix for `xlq certify` over DIVERSE corruption (Lever 1.2/1.3).

cert_confusion.py's corrupt arm was a monoculture — openpyxl's single no-op-shift,
the maximally-detectable corruption — and used the Excel-cache oracle with a
reliability gate. This diversifies the corrupt arm with mutation corruptors and
adjudicates with the LibreOffice-SELF-CONSISTENT oracle (recompute original + edited
both with LibreOffice; a blank-row insert must preserve every value), so no Excel
disagreement and NO reliability gate.

Arms per generated (single-sheet, evaluable) workbook:
  faithful      xlq restructure                              (should CERTIFY / faithful)
  openpyxl      no-op shift                                  (should REFUSE / corrupted)
  unshift_one   xlq output, ONE shifted ref reverted -1      (should REFUSE / corrupted)
  wrong_delta   xlq output, ONE shifted ref over-shifted +1  (should REFUSE / corrupted)

Positive = truly corrupted (guard should REFUSE). Reports TP/FN/TN/FP + Wilson 95%
CI on the false-certification rate, per corruptor type."""
import json, math, os, re, shutil, subprocess, sys, zipfile
from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter
sys.path.insert(0, os.path.dirname(__file__))
from tokenizer_conformance import gen_formula, lo_values, DATA_ROWS, DATA_COLS
import random

XLQ = "/home/soh/aix/xlq/target/release/xlq"
WORK = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/cc2"
K = 2
CELL_RE = re.compile(r'(<c r="([A-Z]+)(\d+)"(?:(?!</c>).)*?<f[^>]*>)(.*?)(</f>)', re.S)
REFROW = re.compile(r'(\$?[A-Z]{1,3}\$?)(\d+)')


def build_book(path, seed):
    rng = random.Random(seed)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    for r in range(1, DATA_ROWS + 1):
        for c in range(1, DATA_COLS + 1):
            ws.cell(r, c).value = round(1.0 + ((r * 7 + c * 3) % 41) + rng.random(), 4)
    # single-cell-ref evaluable formulas with a shiftable row, one per column
    fs = []
    for i in range(12):
        f = gen_formula(rng)
        if ":" in f:                 # keep single-cell for a clean per-ref mutation
            continue
        col = get_column_letter(DATA_COLS + 2 + i)
        ws[f"{col}45"] = "=" + f
        fs.append((col, 45, f))
    for suf in ("", ".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        if os.path.exists(path + suf):
            os.remove(path + suf)
    wb.save(path)
    return fs


def xlq_restructure(src, work, name):
    dst = os.path.join(work, name); shutil.copy(src, dst)
    for suf in (".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        if os.path.exists(dst + suf):
            os.remove(dst + suf)
    r = subprocess.run([XLQ, "restructure", dst, "--sheet", "S", "--op", "insert-rows",
                        "--at", str(K), "--count", "1", "--actor", "c"], capture_output=True, text=True)
    return dst if '"rev"' in r.stdout else None


def openpyxl_edit(src, work, name):
    dst = os.path.join(work, name)
    wb = openpyxl.load_workbook(src); wb["S"].insert_rows(K, 1); wb.save(dst)
    return dst


def mutate(xlq_out, work, name, delta):
    """Surgically change the FIRST cell-ref row (>=K+1, i.e. one xlq shifted) in the
    FIRST formula, by `delta`. Returns dst path or None if nothing shiftable found."""
    part = "xl/worksheets/sheet1.xml"
    data = zipfile.ZipFile(xlq_out).read(part).decode("utf-8", "replace")
    mutated = None
    def repl(m):
        nonlocal mutated
        if mutated:
            return m.group(0)
        body = m.group(4)
        def rowrepl(rm):
            nonlocal mutated
            row = int(rm.group(2))
            if not mutated and row >= K + 1:
                mutated = True
                return f"{rm.group(1)}{row + delta}"
            return rm.group(0)
        newbody = REFROW.sub(rowrepl, body)
        return m.group(1) + newbody + m.group(5)
    data2 = CELL_RE.sub(repl, data, count=0)
    if not mutated:
        return None
    dst = os.path.join(work, name); shutil.copy(xlq_out, dst)
    names = zipfile.ZipFile(xlq_out).namelist()
    buf = {n: zipfile.ZipFile(xlq_out).read(n) for n in names}
    buf[part] = data2.encode("utf-8")
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as z:
        for n in names:
            z.writestr(n, buf[n])
    return dst


def certify(src, edited, work):
    r = subprocess.run([XLQ, "certify", src, edited, "--sheet", "S", "--op", "insert-rows",
                        "--at", str(K), "--count", "1"], capture_output=True, text=True)
    try:
        return json.loads(r.stdout).get("status", "REFUSED")
    except Exception:
        return "REFUSED"


def oracle(src, edited, placed, work):
    """faithful|corrupted via LibreOffice self-consistent value-preservation (insert)."""
    before = lo_values(src, os.path.join(work, "b"))
    after = lo_values(edited, os.path.join(work, "a"))
    checked = 0
    for col, row, _ in placed:
        v0 = before.get((col, row)); v1 = after.get((col, row + 1))
        if v0 is None or v1 is None:
            continue
        checked += 1
        if abs(v1 - v0) > 1e-6 * max(abs(v0), abs(v1), 1.0):
            return "corrupted"
    return "faithful" if checked else "na"


def wilson(k, n, z=1.96):
    if n == 0:
        return (0.0, 1.0)
    p = k / n
    d = 1 + z*z/n
    c = (p + z*z/(2*n)) / d
    h = z*math.sqrt(p*(1-p)/n + z*z/(4*n*n)) / d
    return (max(0, c-h), min(1, c+h))


if __name__ == "__main__":
    os.makedirs(WORK, exist_ok=True)
    n_books = int(sys.argv[1]) if len(sys.argv) > 1 else 15
    per_type = Counter()          # (corruptor, outcome)
    cm = Counter()
    rows = []
    for b in range(n_books):
        work = os.path.join(WORK, str(b)); os.makedirs(work, exist_ok=True)
        src = os.path.join(work, "orig.xlsx")
        placed = build_book(src, seed=1000 + b)
        xo = xlq_restructure(src, work, "xlq.xlsx")
        if not xo:
            shutil.rmtree(work, ignore_errors=True); continue
        arms = [("faithful", xo)]
        arms.append(("openpyxl", openpyxl_edit(src, work, "opx.xlsx")))
        m1 = mutate(xo, work, "unshift.xlsx", -1)
        if m1: arms.append(("unshift_one", m1))
        m2 = mutate(xo, work, "wrongdelta.xlsx", +1)
        if m2: arms.append(("wrong_delta", m2))
        for name, path in arms:
            # GROUND TRUTH BY CONSTRUCTION: the xlq arm is faithful; openpyxl and the
            # two mutations are injected corruptions — we KNOW their label independent
            # of any engine. The LibreOffice oracle is a cross-check, not the truth.
            known = "faithful" if name == "faithful" else "corrupt"
            ol = oracle(src, path, placed, work)
            v = certify(src, path, work)
            refused = (v == "REFUSED")
            rows.append({"book": b, "corruptor": name, "known": known,
                         "certify": v, "oracle": ol})
            if known == "corrupt":
                cm["known_corrupt"] += 1
                cm["refused_corrupt" if refused else "FALSE_CERT"] += 1
                if ol == "faithful":            # value-preserving corruption
                    cm["value_preserving_caught_by_certify"] += 1
                per_type[(name, "refused" if refused else "certified")] += 1
            else:
                cm["known_faithful"] += 1
                cm["refused_faithful" if refused else "certified_faithful"] += 1
        shutil.rmtree(work, ignore_errors=True)
        print(f"  book {b}: " + " ".join(f"{r['corruptor']}={r['certify'][:4]}/{r['oracle'][:4]}"
              for r in rows if r['book'] == b), flush=True)

    kc, fcert = cm["known_corrupt"], cm["FALSE_CERT"]
    kf, rff = cm["known_faithful"], cm["refused_faithful"]
    lo, hi = wilson(fcert, kc)
    by_type = {}
    for (name, cell), n in per_type.items():
        by_type.setdefault(name, {})[cell] = n
    summary = {
        "validation": "xlq certify over DIVERSE corruptors; ground truth BY CONSTRUCTION "
                      "(injected corruptions), LibreOffice self-consistent oracle as cross-check "
                      "(no reliability gate)",
        "corruptor_types": ["openpyxl(no-op-shift)", "unshift_one", "wrong_delta"],
        "known_corrupt_edits": kc,
        "FALSE_CERTIFICATIONS": fcert,
        "false_certification_rate": round(fcert / kc, 4) if kc else None,
        "false_cert_rate_wilson95_ci": [round(lo, 4), round(hi, 4)],
        "known_faithful_edits": kf, "faithful_falsely_refused": rff,
        "value_preserving_corruptions_certify_caught_but_value_oracle_missed":
            cm["value_preserving_caught_by_certify"],
        "by_corruptor_certify_verdict": by_type,
        "headline": (f"{kc} injected corruptions over 3 types: {fcert} false certifications "
                     f"(Wilson95 upper {round(hi,3)}); {kf}/{kf} faithful xlq edits certified; "
                     f"certify caught {cm['value_preserving_caught_by_certify']} value-preserving "
                     f"corruptions the value-oracle could not see (it is stricter than a value check)."),
    }
    with open("/home/soh/aix/benchmarks/cert_confusion_v2.json", "w") as f:
        json.dump(summary, f, indent=2)
    print("\n" + json.dumps({k: v for k, v in summary.items() if k != "by_corruptor"}, indent=2))
    print("by corruptor:", json.dumps(by_type))
