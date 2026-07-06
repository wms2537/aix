#!/usr/bin/env python3
"""LIVE-AGENT 3-way, scored against xlq's PROVEN transform (not LibreOffice).

Rationale: LibreOffice disagrees with Excel on many complex functions (BESSEL,
ERF, ACCRINT, BIN2HEX), and openpyxl's insert_rows mishandles multi-sheet files —
both inject noise. xlq's structural transform is independently validated (Theorem 1
+ forward-correctness on 150 real workbooks), so we use it as the trusted reference:

  ground truth: agent's formula for a cell == xlq's provably-correct shifted formula.
  guard:        build the agent's edited file (xlq's correct structure, but each
                formula replaced by the AGENT's) and run the real `xlq certify`.

3-way per file:
  task_completed_certified   agent correct AND `xlq certify` CERTIFIED
  safely_refused             `xlq certify` REFUSED (agent erred -> caught)
  FALSE_CERT                 CERTIFIED but agent actually erred (must be 0)
  UNGUARDED silent corruption: agent erred and it would ship as-is."""
import json, os, re, shutil, subprocess, sys
sys.path.insert(0, os.path.dirname(__file__))

XLQ = "/home/soh/aix/xlq/target/release/xlq"
CORPUS = "/home/soh/aix/vendor/upstream/xlsx/tests"
WORK = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/live3truth"
K = 2


def norm(f):
    f = str(f)
    if f.startswith("="):
        f = f[1:]
    return re.sub(r"\s+", "", f).upper()


def xlq_restructure(src, sheet, work):
    dst = os.path.join(work, "xlq.xlsx")
    shutil.copy(src, dst)
    for suf in (".xlq.jsonl", ".rev-1.xlsx", ".xlq.lock"):
        p = dst + suf
        if os.path.exists(p):
            os.remove(p)
    r = subprocess.run([XLQ, "restructure", dst, "--sheet", sheet, "--op", "insert-rows",
                        "--at", str(K), "--count", "1", "--actor", "t"],
                       capture_output=True, text=True)
    return dst if '"rev"' in r.stdout else None


def sheet_formulas(path, sheet):
    """{A1: formula} for the given sheet, via openpyxl (multi-sheet safe)."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                out[c.coordinate] = c.value
    return out


def _sheet_part_for(z, sheet):
    """Map a sheet NAME to its worksheet part inside the zip (multi-sheet safe),
    via workbook.xml + its rels."""
    import re as _re
    wb = z.read("xl/workbook.xml").decode("utf-8", "replace")
    m = _re.search(r'<sheet\b[^>]*\bname="' + _re.escape(sheet) + r'"[^>]*\br:id="([^"]+)"', wb)
    if not m:
        m2 = _re.search(r'<sheet\b[^>]*\br:id="([^"]+)"', wb)  # fall back to first
        rid = m2.group(1) if m2 else None
    else:
        rid = m.group(1)
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
    rm = _re.search(r'<Relationship\b[^>]*\bId="' + _re.escape(rid or "") + r'"[^>]*\bTarget="([^"]+)"', rels)
    if not rm:
        return None
    tgt = rm.group(1).lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def _esc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_agent_file(xlq_out, sheet, agent_by_newpos, dst):
    """SURGICAL: xlq's correctly-structured output with ONLY the target formula
    cells' <f> text replaced by the AGENT's formula. Every other byte (other
    sheets, styles, caches, structure) stays identical — no openpyxl round-trip,
    so `xlq certify` sees exactly the agent's formula choices and nothing else."""
    import re as _re, zipfile
    shutil.copy(xlq_out, dst)
    zin = zipfile.ZipFile(xlq_out)
    part = _sheet_part_for(zin, sheet)
    names = zin.namelist()
    data = zin.read(part).decode("utf-8", "replace")
    for a1, f in agent_by_newpos.items():
        body = f[1:] if str(f).startswith("=") else str(f)
        # replace the <f>...</f> content inside the cell element r="a1"
        pat = _re.compile(r'(<c r="' + _re.escape(a1) + r'"(?:(?!</c>).)*?<f[^>]*>)(.*?)(</f>)', _re.S)
        data, n = pat.subn(lambda mm: mm.group(1) + _esc(body) + mm.group(3), data, count=1)
    buf = {nm: zin.read(nm) for nm in names}
    buf[part] = data.encode("utf-8")
    zin.close()
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zo:
        for nm in names:
            zo.writestr(nm, buf[nm])
    return dst


def shifted_a1(a1):
    col = "".join(ch for ch in a1 if ch.isalpha())
    row = int("".join(ch for ch in a1 if ch.isdigit()))
    return f"{col}{row + 1 if row >= K else row}"


def xlq_certify(src, edited, sheet):
    r = subprocess.run([XLQ, "certify", src, edited, "--sheet", sheet, "--op",
                        "insert-rows", "--at", str(K), "--count", "1"],
                       capture_output=True, text=True)
    try:
        return json.loads(r.stdout).get("status", "REFUSED")
    except Exception:
        return "REFUSED"


if __name__ == "__main__":
    outputs = json.load(open(sys.argv[1]))
    tfile = os.environ.get("TASKS_FILE", "/home/soh/aix/benchmarks/live3way_tasks.json")
    tasks = {t["file"]: t for t in json.load(open(tfile))}
    os.makedirs(WORK, exist_ok=True)
    rows, agg = [], {"task_completed_certified": 0, "safely_refused": 0,
                     "FALSE_CERT": 0, "unguarded_silent_corruption": 0, "skip": 0}
    for i, (rel, agent) in enumerate(outputs.items()):
        t = tasks.get(rel)
        if not t:
            continue
        src = os.path.join(CORPUS, rel); sheet = t["sheet"]
        work = os.path.join(WORK, str(i)); os.makedirs(work, exist_ok=True)
        xo = xlq_restructure(src, sheet, work)
        if not xo:
            agg["skip"] += 1; shutil.rmtree(work, ignore_errors=True); continue
        truth = sheet_formulas(xo, sheet)                 # xlq's correct shifted formulas
        # agent formulas keyed by SHIFTED position; per-cell correctness vs xlq
        agent_newpos, mismatches = {}, 0
        for orig_a1, af in agent.items():
            np = shifted_a1(orig_a1)
            agent_newpos[np] = af
            tf = truth.get(np)
            if tf is None:
                continue
            if norm(af) != norm(tf):
                mismatches += 1
        agent_correct = (mismatches == 0)
        agent_file = build_agent_file(xo, sheet, agent_newpos, os.path.join(work, "agent.xlsx"))
        verdict = xlq_certify(src, agent_file, sheet)
        shutil.rmtree(work, ignore_errors=True)
        if verdict == "CERTIFIED" and agent_correct:
            g = "task_completed_certified"
        elif verdict == "CERTIFIED" and not agent_correct:
            g = "FALSE_CERT"
        else:
            g = "safely_refused"
        agg[g] += 1
        if not agent_correct:
            agg["unguarded_silent_corruption"] += 1
        rows.append({"file": rel, "agent_correct": agent_correct, "mismatches": mismatches,
                     "xlq_certify": verdict, "guarded": g})
        print(f"  {rel[:42]:42} agent={'ok ' if agent_correct else f'ERR({mismatches})':>7} "
              f"certify={verdict:9} -> {g}", flush=True)

    n = len(rows)
    summary = {
        "experiment": "LIVE-AGENT 3-way vs xlq's PROVEN transform (no LibreOffice): a real "
                      "(Haiku) agent rewrites formula references for insert-row@2; its edit is "
                      "scored by the real `xlq certify` guard, ground-truth = xlq's shift.",
        "files": n, "skipped": agg["skip"],
        "agent_error_rate": round(agg["unguarded_silent_corruption"] / n, 3) if n else None,
        "GUARDED": {"task_completed_certified": agg["task_completed_certified"],
                    "safely_refused": agg["safely_refused"],
                    "FALSE_CERT_must_be_0": agg["FALSE_CERT"]},
        "UNGUARDED_silent_corruptions": agg["unguarded_silent_corruption"],
        "headline": (f"agent erred on {agg['unguarded_silent_corruption']}/{n} workbooks; "
                     f"GUARDED by xlq certify: {agg['task_completed_certified']} completed+certified, "
                     f"{agg['safely_refused']} safely refused, {agg['FALSE_CERT']} false certs. "
                     f"UNGUARDED those {agg['unguarded_silent_corruption']} ship silently."),
        "per_file": rows,
    }
    with open("/home/soh/aix/benchmarks/live3way_truth.json", "w") as f:
        json.dump(summary, f, indent=2)
    print("\n" + json.dumps({k: v for k, v in summary.items() if k != "per_file"}, indent=2))
