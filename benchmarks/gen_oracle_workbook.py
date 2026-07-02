#!/usr/bin/env python3
"""Generate the differential-oracle workbook for ironcalc vs LibreOffice.

Reads benchmarks/oracle-cases.json, writes <outdir>/oracle.xlsx via openpyxl:
  - shared data block on sheet "T" (columns A-E, see _meta in the JSON)
  - one case per row, formula in column G, starting at G1
  - <outdir>/manifest.json mapping row -> {function, formula}

openpyxl writes NO cached formula values — that is the point: when
LibreOffice converts the file (soffice --headless --convert-to xlsx) it must
compute every formula itself, and the computed <v> values in the converted
file become the LibreOffice side of the oracle.

Because the file is OOXML, post-2007 function names must be stored with
Excel's `_xlfn.` (or `_xlfn._xlws.`) prefix or LibreOffice imports them as
unknown macro calls (#NAME?). The prefixing below only rewrites how the
formula is STORED; the case formulas themselves stay canonical Excel-UI
spellings (and are fed verbatim to ironcalc by oracle-compare).

Usage: gen_oracle_workbook.py <oracle-cases.json> <outdir>
"""

import json
import os
import sys

import openpyxl

# Functions introduced after Excel 2007, stored in OOXML with the `_xlfn.`
# prefix (source: Microsoft "Excel functions that are stored with prefixes"
# / [MS-XLSX] future-function list), restricted to names that appear in the
# case table plus close relatives.
XLFN = {
    "ACOT", "ACOTH", "AGGREGATE", "ARABIC", "ARRAYTOTEXT", "BASE",
    "BETA.DIST", "BETA.INV", "BINOM.DIST", "BINOM.DIST.RANGE", "BINOM.INV",
    "BITAND", "BITLSHIFT", "BITOR", "BITRSHIFT", "BITXOR", "BYCOL", "BYROW",
    "CEILING.MATH", "CEILING.PRECISE", "CHISQ.DIST", "CHISQ.DIST.RT",
    "CHISQ.INV", "CHISQ.INV.RT", "CHISQ.TEST", "CHOOSECOLS", "CHOOSEROWS",
    "COMBINA", "CONCAT", "CONFIDENCE.NORM", "CONFIDENCE.T", "COT", "COTH",
    "COVARIANCE.P", "COVARIANCE.S", "CSC", "CSCH", "DAYS", "DECIMAL", "DROP",
    "ERF.PRECISE", "ERFC.PRECISE", "EXPAND", "EXPON.DIST", "F.DIST",
    "F.DIST.RT", "F.INV", "F.INV.RT", "F.TEST", "FLOOR.MATH", "FLOOR.PRECISE",
    "FORECAST.LINEAR", "FORMULATEXT", "GAMMA", "GAMMA.DIST", "GAMMA.INV",
    "GAMMALN.PRECISE", "GAUSS", "HSTACK", "HYPGEOM.DIST", "IFNA", "IFS",
    "ISFORMULA", "ISOMITTED", "ISOWEEKNUM", "LAMBDA",
    "LET", "LOGNORM.DIST", "LOGNORM.INV", "MAKEARRAY", "MAP", "MAXIFS",
    "MINIFS", "MODE.MULT", "MODE.SNGL", "MUNIT", "NEGBINOM.DIST",
    "NORM.DIST", "NORM.INV", "NORM.S.DIST", "NORM.S.INV",
    "NUMBERVALUE", "PDURATION", "PERCENTILE.EXC", "PERCENTILE.INC",
    "PERCENTOF", "PERCENTRANK.EXC", "PERCENTRANK.INC", "PERMUTATIONA", "PHI",
    "POISSON.DIST", "QUARTILE.EXC", "QUARTILE.INC", "RANDARRAY", "RANK.AVG",
    "RANK.EQ", "REDUCE", "REGEXEXTRACT", "REGEXREPLACE", "REGEXTEST", "RRI",
    "SCAN", "SEC", "SECH", "SEQUENCE", "SHEET", "SHEETS", "SKEW.P", "SORTBY",
    "STDEV.P", "STDEV.S", "SWITCH", "T.DIST", "T.DIST.2T", "T.DIST.RT",
    "T.INV", "T.INV.2T", "T.TEST", "TAKE", "TEXTAFTER", "TEXTBEFORE",
    "TEXTJOIN", "TEXTSPLIT", "TOCOL", "TOROW", "TRIMRANGE", "UNICHAR",
    "UNICODE", "UNIQUE", "VALUETOTEXT", "VAR.P", "VAR.S", "VSTACK",
    "WEIBULL.DIST", "WRAPCOLS", "WRAPROWS", "XLOOKUP",
    "XMATCH", "XOR", "Z.TEST",
}
# These two are stored with the `_xlfn._xlws.` prefix.
XLWS = {"FILTER", "SORT"}
# LibreOffice 24.8 OOXML-import quirk (verified empirically on this machine):
# these post-2007 names are only recognized UNPREFIXED — with the Excel
# `_xlfn.` storage prefix LO leaves them as unknown macro calls (#NAME?).
# Since the workbook exists solely for LibreOffice to compute, they are
# stored unprefixed here (Excel storage convention is irrelevant for it):
#   IMCOSH IMCOT IMCSC IMCSCH IMSEC IMSECH IMSINH IMTAN
#   NETWORKDAYS.INTL WORKDAY.INTL ISO.CEILING

IDENT_START = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz_")
IDENT_CONT = IDENT_START | set("0123456789.")


def prefix_formula(formula: str) -> str:
    """Rewrite `NAME(` tokens (outside string literals) with OOXML storage
    prefixes. Only names in XLFN/XLWS are touched."""
    out = []
    i, n = 0, len(formula)
    while i < n:
        ch = formula[i]
        if ch == '"':  # string literal, "" is the escape
            j = i + 1
            while j < n:
                if formula[j] == '"':
                    if j + 1 < n and formula[j + 1] == '"':
                        j += 2
                        continue
                    j += 1
                    break
                j += 1
            out.append(formula[i:j])
            i = j
            continue
        if ch in IDENT_START:
            j = i + 1
            while j < n and formula[j] in IDENT_CONT:
                j += 1
            name = formula[i:j]
            k = j
            while k < n and formula[k] == ' ':
                k += 1
            upper = name.upper()
            if k < n and formula[k] == '(':
                if upper in XLWS:
                    out.append("_xlfn._xlws." + upper)
                elif upper in XLFN:
                    out.append("_xlfn." + upper)
                else:
                    out.append(name)
            else:
                out.append(name)
            i = j
            continue
        out.append(ch)
        i += 1
    return "".join(out)


def write_data_block(ws) -> None:
    a = [2, 4, 6, 8, 10, -3, 0, 7.5, 100, 1]
    b = ["alpha", "Beta", "gamma DELTA", "2026-03-15", "x,y;z",
         " padded ", "", "MiXeD", "100", "-5"]
    for i, v in enumerate(a, start=1):
        ws.cell(row=i, column=1).value = v
    for i, v in enumerate(b, start=1):
        if v == "":
            continue  # B7: blank cell (see _meta note in oracle-cases.json)
        c = ws.cell(row=i, column=2)
        c.value = v
        c.data_type = "s"  # force text ("2026-03-15", "100", "-5" stay text)
    for i, v in enumerate([1, 2, 3, 4, 5], start=1):
        ws.cell(row=i, column=3).value = v
    for i, v in enumerate([10, 20, 30, 40, 50], start=1):
        ws.cell(row=i, column=4).value = v
    ws.cell(row=1, column=5).value = True
    ws.cell(row=2, column=5).value = False


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        return 2
    cases_path, outdir = sys.argv[1], sys.argv[2]
    os.makedirs(outdir, exist_ok=True)

    with open(cases_path) as f:
        table = json.load(f)
    table.pop("_meta", None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    write_data_block(ws)

    manifest = {}
    row = 0
    for func in sorted(table.keys()):
        for formula in table[func]:
            row += 1
            ws.cell(row=row, column=7).value = prefix_formula(formula)
            manifest[str(row)] = {"function": func, "formula": formula}

    xlsx_path = os.path.join(outdir, "oracle.xlsx")
    wb.save(xlsx_path)
    with open(os.path.join(outdir, "manifest.json"), "w") as f:
        json.dump(manifest, f, indent=1, sort_keys=True)
        f.write("\n")
    print(f"wrote {xlsx_path}: {row} cases on sheet T (G1:G{row})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
