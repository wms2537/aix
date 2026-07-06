#!/usr/bin/env python3
"""Score the LIVE-AGENT 3-way slice. Given each agent's corrected formulas, build
the agent-authored edited file, run it through the certify-or-refuse guard, and
label it with the INDEPENDENT LibreOffice oracle. Three outcomes per file:

  GUARDED path (agent edit -> certifier):
    task_completed_certified   agent was right AND certifier CERTIFIED  (win)
    safely_refused             certifier REFUSED                        (no silent corruption)
    silently_corrupted_GUARDED certifier CERTIFIED but oracle says WRONG (FALSE CERT — must be 0)
  UNGUARDED path (agent edit committed as-is):
    silently_corrupted_UNGUARDED  agent erred and it would ship silently (the risk the guard removes)

Usage: live3way_score.py <agent_outputs.json> [certifier]  (certifier: python|xlq)"""
import json, os, shutil, subprocess, sys
sys.path.insert(0, os.path.dirname(__file__))
from foreign_certify import certify_foreign
from forward_correctness import orig_formula_caches, lo_grid, check, XLQ

CORPUS = "/home/soh/aix/vendor/upstream/xlsx/tests"
WORK = "/tmp/claude-1000/-home-soh-aix/a1b7f99e-cc58-4254-b95a-10d56f89029d/scratchpad/live3way"
K = 2


def build_agent_edit(orig_path, sheet, corrected, dst):
    import openpyxl
    wb = openpyxl.load_workbook(orig_path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    ws.insert_rows(K, 1)                       # positions shift; formulas left unshifted
    for a1, newf in corrected.items():
        col = "".join(ch for ch in a1 if ch.isalpha())
        row = int("".join(ch for ch in a1 if ch.isdigit()))
        newrow = row + 1 if row >= K else row
        ws[f"{col}{newrow}"] = newf if str(newf).startswith("=") else "=" + str(newf)
    wb.save(dst)
    return dst


def oracle_label(orig_path, edited_path, work):
    """faithful | corrupted | na, by LibreOffice recompute vs original Excel cache."""
    caches = orig_formula_caches(orig_path)
    if not caches:
        return "na"
    grid = lo_grid(edited_path, work)
    checked, matched, _ = check(caches, grid)
    if checked == 0:
        return "na"
    return "faithful" if matched == checked else "corrupted"


def certify(orig_path, edited_path, sheet, which):
    if which == "xlq":
        r = subprocess.run([XLQ, "certify", orig_path, edited_path, "--sheet", sheet,
                            "--op", "insert-rows", "--at", str(K), "--count", "1"],
                           capture_output=True, text=True)
        try:
            return json.loads(r.stdout).get("status", "REFUSED")
        except Exception:
            return "REFUSED"
    res = certify_foreign(orig_path, edited_path, K)
    if res is None:
        return "REFUSED"
    return res if isinstance(res, str) else res[0]


if __name__ == "__main__":
    outputs = json.load(open(sys.argv[1]))                 # {file: {a1: corrected_formula}}
    which = sys.argv[2] if len(sys.argv) > 2 else "python"
    tasks = {t["file"]: t for t in json.load(open("/home/soh/aix/benchmarks/live3way_tasks.json"))}
    os.makedirs(WORK, exist_ok=True)

    rows, agg = [], {"task_completed_certified": 0, "safely_refused": 0,
                     "silently_corrupted_GUARDED_false_cert": 0,
                     "silently_corrupted_UNGUARDED": 0, "na": 0}
    for i, (rel, corrected) in enumerate(outputs.items()):
        t = tasks.get(rel)
        if not t:
            continue
        src = os.path.join(CORPUS, rel)
        work = os.path.join(WORK, str(i)); os.makedirs(work, exist_ok=True)
        dst = os.path.join(work, "agent.xlsx")
        try:
            build_agent_edit(src, t["sheet"], corrected, dst)
        except Exception as e:
            print(f"  ! build {rel}: {e}"); shutil.rmtree(work, ignore_errors=True); continue
        olabel = oracle_label(src, dst, work)
        verdict = certify(src, dst, t["sheet"], which)
        shutil.rmtree(work, ignore_errors=True)
        if olabel == "na":
            agg["na"] += 1
            rows.append({"file": rel, "oracle": "na", "certifier": verdict}); continue
        agent_correct = (olabel == "faithful")
        # GUARDED 3-way
        if verdict == "CERTIFIED" and agent_correct:
            g = "task_completed_certified"
        elif verdict == "CERTIFIED" and not agent_correct:
            g = "silently_corrupted_GUARDED_false_cert"
        else:
            g = "safely_refused"
        agg[g] += 1
        # UNGUARDED: what ships if committed as-is
        if not agent_correct:
            agg["silently_corrupted_UNGUARDED"] += 1
        rows.append({"file": rel, "oracle": olabel, "certifier": verdict, "guarded_outcome": g})
        print(f"  {rel[:44]:44} agent={'ok ' if agent_correct else 'ERR'} "
              f"certifier={verdict:9} -> {g}", flush=True)

    scored = sum(1 for r in rows if r["oracle"] != "na")
    ug_corrupt = agg["silently_corrupted_UNGUARDED"]
    summary = {
        "experiment": f"LIVE-AGENT 3-way (certifier={which}): a real agent rewrites formula "
                      "references for insert-row@2; its OWN edit is scored by the certify-or-"
                      "refuse guard + INDEPENDENT LibreOffice oracle",
        "files_scored": scored, "na": agg["na"],
        "GUARDED": {"task_completed_certified": agg["task_completed_certified"],
                    "safely_refused": agg["safely_refused"],
                    "silently_corrupted_FALSE_CERT": agg["silently_corrupted_GUARDED_false_cert"]},
        "UNGUARDED_would_silently_corrupt": ug_corrupt,
        "agent_error_rate": round(ug_corrupt / scored, 3) if scored else None,
        "headline": (f"agent erred on {ug_corrupt}/{scored}; GUARDED: "
                     f"{agg['task_completed_certified']} completed+certified, "
                     f"{agg['safely_refused']} safely refused, "
                     f"{agg['silently_corrupted_GUARDED_false_cert']} silent corruptions. "
                     f"UNGUARDED those {ug_corrupt} errors ship silently."),
        "per_file": rows,
    }
    with open(f"/home/soh/aix/benchmarks/live3way_{which}.json", "w") as f:
        json.dump(summary, f, indent=2)
    print("\n" + json.dumps({k: v for k, v in summary.items() if k != "per_file"}, indent=2))
