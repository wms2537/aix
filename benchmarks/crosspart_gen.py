#!/usr/bin/env python3
"""Generate a CROSS-PART fixture corpus: workbooks that deliberately exercise
every reference-bearing OOXML part the uniform σ algorithm claims to shift —
charts, cross-sheet references, defined names, conditional formatting, data
validation, merged cells — so the contribution is validated where it is novel,
not only on single-sheet flat grids. Each fixture records its reference
inventory as JSON so the verifier can compute expected post-edit targets."""
import json, os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/home/soh/aix/fixtures/crosspart"
os.makedirs(OUT, exist_ok=True)


def build(idx, ndata, chart_rows, xref_target_row, cf_range, dv_range, defined_row):
    """One fixture: Data sheet with numeric columns + a chart over a column,
    a CF rule, a DV, a merged region, a defined name into Data, and a Report
    sheet cross-referencing Data. Returns the reference inventory."""
    wb = openpyxl.Workbook()
    d = wb.active
    d.title = "Data"
    d["A1"] = "k"
    d["B1"] = "v"
    for r in range(2, 2 + ndata):
        d.cell(row=r, column=1, value=r - 1)
        d.cell(row=r, column=2, value=(r - 1) * 10)
    sum_row = 2 + ndata
    d.cell(row=sum_row, column=2, value=f"=SUM(B2:B{sum_row - 1})")  # straddles inserts
    d.cell(row=sum_row + 1, column=2, value=f"=B{xref_target_row}*2")  # single ref
    d.cell(row=sum_row + 2, column=4, value=f"=$B${sum_row}")           # absolute
    d.merge_cells(f"D{cf_range[0]}:E{cf_range[0]}")                     # merged region
    # a bar chart over B1:B(chart_rows)
    chart = BarChart()
    data = Reference(d, min_col=2, min_row=1, max_row=chart_rows)
    chart.add_data(data, titles_from_data=True)
    d.add_chart(chart, "G2")
    # conditional formatting over a column range
    d.conditional_formatting.add(
        f"B{cf_range[0]}:B{cf_range[1]}",
        CellIsRule(operator="greaterThan", formula=[f"$B${cf_range[0]}"], fill=None),
    )
    # data validation referencing a column
    dv = DataValidation(type="list", formula1=f"$A${dv_range[0]}:$A${dv_range[1]}")
    d.add_data_validation(dv)
    dv.add(f"F{dv_range[0]}:F{dv_range[1]}")
    # cross-sheet reference from Report into Data
    rep = wb.create_sheet("Report")
    rep["A1"] = f"=Data!B{sum_row}"
    rep["A2"] = f"=SUM(Data!B2:B{sum_row - 1})"
    # a defined name into Data
    wb.defined_names.add(DefinedName("Total", attr_text=f"Data!$B${sum_row}"))
    path = os.path.join(OUT, f"cp{idx:02d}.xlsx")
    wb.save(path)
    return {
        "file": f"cp{idx:02d}.xlsx",
        "edited_sheet": "Data",
        "ndata": ndata,
        "references": {
            # (part, kind, original A1 body on the Data sheet's row/col axis)
            "insheet_sum": {"cell": f"B{sum_row}", "ref": f"B2:B{sum_row - 1}"},
            "insheet_single": {"cell": f"B{sum_row + 1}", "ref": f"B{xref_target_row}"},
            "insheet_abs": {"cell": f"D{sum_row + 2}", "ref": f"$B${sum_row}"},
            "chart": {"part": "chart", "ref": f"Data!$B$2:$B${chart_rows}"},
            "cf": {"part": "cf", "sqref": f"B{cf_range[0]}:B{cf_range[1]}", "formula": f"$B${cf_range[0]}"},
            "dv": {"part": "dv", "sqref": f"F{dv_range[0]}:F{dv_range[1]}", "formula": f"$A${dv_range[0]}:$A${dv_range[1]}"},
            "merged": {"part": "merged", "ref": f"D{cf_range[0]}:E{cf_range[0]}"},
            "xref_single": {"part": "Report", "cell": "A1", "ref": f"Data!B{sum_row}"},
            "xref_range": {"part": "Report", "cell": "A2", "ref": f"Data!B2:B{sum_row - 1}"},
            "defined": {"part": "workbook", "ref": f"Data!$B${sum_row}"},
        },
        "sum_row": sum_row,
    }


if __name__ == "__main__":
    inventory = []
    # vary the geometry so edits land above / inside / below the referenced ranges
    configs = [
        (6, 6, 4, (3, 6), (2, 5), 4),
        (8, 8, 5, (4, 8), (2, 7), 6),
        (5, 5, 3, (2, 5), (3, 5), 3),
        (10, 10, 6, (5, 10), (2, 9), 7),
        (7, 7, 7, (3, 7), (2, 6), 5),
        (12, 12, 8, (6, 12), (4, 11), 9),
        (4, 4, 2, (2, 4), (2, 4), 2),
        (9, 9, 4, (3, 9), (5, 9), 8),
        (6, 6, 5, (4, 6), (2, 6), 3),
        (11, 11, 9, (2, 11), (3, 10), 10),
    ]
    for i, c in enumerate(configs, 1):
        inventory.append(build(i, *c))
    with open(os.path.join(OUT, "inventory.json"), "w") as f:
        json.dump(inventory, f, indent=2)
    print(f"generated {len(inventory)} cross-part fixtures in {OUT}")
    # quick sanity: chart + xref + defined-name present in one
    import zipfile
    z = zipfile.ZipFile(os.path.join(OUT, "cp01.xlsx"))
    print("cp01 parts:", [n for n in z.namelist() if any(k in n for k in ("chart", "worksheets", "workbook.xml"))])
